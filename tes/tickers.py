"""
extract_idx_tickers.py
----------------------
Ekstrak semua ticker (Kode Saham) dari file idx_summary.xlsx
dan format sebagai IDX_TICKERS_RAW string siap pakai.

Usage:
    python extract_idx_tickers.py
    python extract_idx_tickers.py --xlsx path/to/file.xlsx
    python extract_idx_tickers.py --output tickers.py
"""

import argparse
import textwrap
from pathlib import Path

import openpyxl


XLSX_DEFAULT = Path(__file__).parent / "idx_summary.xlsx"
TICKER_COLUMN = "Kode Saham"
COLS_PER_LINE = 15


def extract_tickers(xlsx_path: str) -> list[str]:
    """Baca semua ticker dari kolom 'Kode Saham' di xlsx."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb.active

    # Cari index kolom Kode Saham dari header row
    header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    try:
        col_idx = list(header).index(TICKER_COLUMN)
    except ValueError:
        raise ValueError(
            f"Kolom '{TICKER_COLUMN}' tidak ditemukan. "
            f"Kolom yang tersedia: {list(header)}"
        )

    tickers = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[col_idx]
        if val and isinstance(val, str) and val.strip():
            tickers.append(val.strip())

    wb.close()
    return sorted(set(tickers))  # deduplicate & sort alphabetically


def format_as_raw_string(tickers: list[str], cols: int = COLS_PER_LINE) -> str:
    """Format list ticker jadi IDX_TICKERS_RAW string (mirip format aslinya)."""
    lines = []
    for i in range(0, len(tickers), cols):
        lines.append(" ".join(tickers[i : i + cols]))

    body = "\n".join(lines)
    return f'IDX_TICKERS_RAW = """\n{body}\n""".split()'


def main():
    parser = argparse.ArgumentParser(description="Extract IDX tickers from xlsx")
    parser.add_argument(
        "--xlsx",
        default=XLSX_DEFAULT,
        help=f"Path ke file xlsx (default: {XLSX_DEFAULT})",
    )
    parser.add_argument(
        "--output",
        default=None,
        help="Simpan hasil ke file .py (opsional, default: print ke terminal)",
    )
    args = parser.parse_args()

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.exists():
        raise FileNotFoundError(f"File tidak ditemukan: {xlsx_path}")

    print(f"📂 Membaca: {xlsx_path}")
    tickers = extract_tickers(str(xlsx_path))
    print(f"✅ Total ticker unik: {len(tickers)}")

    result = format_as_raw_string(tickers)

    if args.output:
        out_path = Path(args.output)
        out_path.write_text(result + "\n", encoding="utf-8")
        print(f"💾 Disimpan ke: {out_path}")
    else:
        print("\n" + "=" * 60)
        print(result)
        print("=" * 60)


if __name__ == "__main__":
    main()